import cv2
import numpy as np
import os
from openpyxl import Workbook,load_workbook

def colorTran(source,target,number):
	sImg = cv2.imread('./images/'+source)
	#sImg = cv2.cvtColor(sImg, cv2.COLOR_BGR2LAB)
	sImg = cv2.cvtColor(sImg, cv2.COLOR_BGR2RGB)
	tImg = cv2.imread('./images/'+target)
	#tImg = cv2.cvtColor(tImg, cv2.COLOR_BGR2LAB)
	tImg = cv2.cvtColor(tImg, cv2.COLOR_BGR2RGB)

	#讀取平均值和標準差
	sMean, sStd = cv2.meanStdDev(sImg)
	tMean, tStd = cv2.meanStdDev(tImg)
	sMean = np.hstack(np.around(sMean, decimals=6))
	sStd = np.hstack(np.around(sStd, decimals=6))
	tMean = np.hstack(np.around(tMean, decimals=6))
	tStd = np.hstack(np.around(tStd, decimals=6))


	#先找到最終transfer完的照片
	height, width, channel = sImg.shape
	trImg = cv2.imread('./images/'+source)
	trImg = cv2.cvtColor(trImg, cv2.COLOR_BGR2RGB)
	for i in range (0,height): 
		for j in range (0,width): 
			for k in range (0,channel): 
				s = sImg[i,j,k] 
				s = (s - sMean[k]) * (tStd[k] / sStd[k]) + tMean[k]
				s = round(s)
				if s < 0:
					s = 0
				if s > 255:
					s = 255  
				trImg[i,j,k] = s
	trMean, trStd = cv2.meanStdDev(trImg)
	trMean = np.hstack(np.around(trMean, decimals=6))
	trStd = np.hstack(np.around(trStd, decimals=6))

	
	#開啟excel檔案
	if number==1:
		wb = Workbook()
		ws = wb.active
		ws.title='Sample Pair'
	elif number==2:
		wb = load_workbook('OWCT-EXCEL.xlsx')
		ws = wb.create_sheet('User Pair')
	ws.append(['Sample Pair','Weight','mean','mean','mean','std','std','std','color distance','color distance','Total','Mark'])
	ws.append(['','','red','green','blue','red','green','blue','source','target','TCD'])

	#從weight=0.00到weight=1.00，一百零一張圖片數據
	w = 0.00
	TCD_MIN=0
	Weight_MIN=0.0
	height, width, channel = sImg.shape
	iImg = cv2.imread('./images/'+source)
	iImg = cv2.cvtColor(iImg, cv2.COLOR_BGR2RGB)

	for n in range(0,101):
		ESI=0
		ETI=0
		TCD=0
		for i in range (0,height): 
			for j in range (0,width): 
				for k in range (0,channel): 
					s = sImg[i,j,k]
					s = ((w*tStd[k]+(1-w)*sStd[k]) / sStd[k]) * (s-sMean[k]) + tMean[k]*w + (1-w)*sMean[k]
					s = round(s)
					if s < 0:
						s = 0
					if s > 255:
						s = 255
					if k==0:
						ESI+=3*(s-sImg[i,j,k])*(s-sImg[i,j,k])
						ETI+=3*(s-trImg[i,j,k])*(s-trImg[i,j,k])
					elif k==1:
						ESI+=4*(s-sImg[i,j,k])*(s-sImg[i,j,k])
						ETI+=4*(s-trImg[i,j,k])*(s-trImg[i,j,k])
					elif k==2:
						ESI+=2*(s-sImg[i,j,k])*(s-sImg[i,j,k])
						ETI+=2*(s-trImg[i,j,k])*(s-trImg[i,j,k])
					iImg[i,j,k]=s
		iMean, iStd = cv2.meanStdDev(iImg)
		iMean = np.hstack(np.around(iMean, decimals=6))
		iStd = np.hstack(np.around(iStd, decimals=6))
		#iImg = cv2.cvtColor(iImg, cv2.COLOR_LAB2BGR)
		#cv2.imwrite('./images/tr'+str(n)+'.bmp',iImg)
		ESI = ESI/(height*width)
		ETI = ETI/(height*width)
		ESI = round(ESI,6)
		ETI = round(ETI,6)
		TCD = abs(ESI - ETI)
		TCD = round(TCD,6)
		print("Weight=",w," TCD=",TCD)
		# print("Weight=",w)
		# print("ESI^2=",ESI)
		# print("ETI^2=",ETI)
		# print("TCD=",TCD)
		# print("Mean=",iMean)
		# print("std=",iStd)
		# print()

		#輸出資料到excel中
		if n==0:
			TCD_MIN=TCD
			ws.append([n, w, iMean[0],iMean[1],iMean[2],iStd[0],iStd[1],iStd[2]])
		elif n==100:
			ws.append([n, w, iMean[0],iMean[1],iMean[2],iStd[0],iStd[1],iStd[2],ESI,ETI,TCD])
		else:
			ws.append([n, w, iMean[0],iMean[1],iMean[2],iStd[0],iStd[1],iStd[2],ESI,ETI,TCD])
	
		if TCD<TCD_MIN:
			Weight_MIN=w
			TCD_MIN=TCD
		w=round(w+0.01,2)
	print("TCD_MIN=",TCD_MIN," Weight_MIN=",Weight_MIN)
	ws.cell(row = Weight_MIN*100+3,column=12).value='Optimal'

	wb.save('OWCT-EXCEL.xlsx')


	#輸出最佳圖片
	for i in range (0,height): 
		for j in range (0,width): 
			for k in range (0,channel): 
				s = sImg[i,j,k] 
				s = ((Weight_MIN*tStd[k]+(1-Weight_MIN)*sStd[k]) / sStd[k]) * (s-sMean[k]) + tMean[k]*Weight_MIN + (1-Weight_MIN)*sMean[k]
				s = round(s)
				if s < 0:
					s = 0
				if s > 255:
					s = 255  
				sImg[i,j,k] = s
	sImg = cv2.cvtColor(sImg,cv2.COLOR_RGB2BGR) 
	cv2.imwrite('./images/OWCT'+str(number)+'-'+str(Weight_MIN)+'.bmp',sImg)
	

files= os.listdir('./images')
sources = ['s1.bmp','s2.bmp']
targets = ['t1.bmp','t2.bmp']
for i in range(2):
	if os.path.isfile('./images/'+sources[i]) and os.path.isfile('./images/'+targets[i]):
		print("第",i+1,'張照片轉換中...\n')
		colorTran(sources[i],targets[i],i+1)
	else:
		print(sources[i]," doesn't exist.")

os.system('pause')
